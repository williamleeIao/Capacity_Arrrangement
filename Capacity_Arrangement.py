import pandas as pd
import re
from openpyxl import load_workbook
import os
from openpyxl import worksheet
import time

full_path = os.path.abspath(os.getcwd())
file_name = full_path + "\\" + "PLANNER_INPUT_FORECAST_HERE.xlsx"
error_file = full_path + "\\" + "error_miss.txt"
# <!------------- Change the template file for different value----------------------->
template_file_io_to_write = full_path + "\\" + "Cap_Ana_Controller_Q2_2020_SO+FC.xlsx"
save_file_name = full_path + "\\" + "PLANNER_INPUT_FORECAST.txt"


def spliiter_successfully_checked(strArray):
    if len(strArray) == 1:
        buildfile = strArray[0]
        target = ''
        return False
    elif len(strArray) == 2:
        buildfile, target = strArray[0], strArray[1]
        return True
    else:
        pass  # handle error; there's two #s


def part_number_process(elements, df, counter_df):
    found = False
    temp_list = []
    regex = r"([A-Z]?\d+)[A-Z]-(\d)"
    # 0. Separate the part number
    z = re.match(regex, df.loc[counter_df][0])
    # result [0] ---> Main Part number
    # result [1] ---> Sub Part Number
    Full_string = z.group(1) + '*-' + z.group(2) + '**L'
    # <1!---------- Adding more and appending at last column  if more column   ---------->
    temp_list = [Full_string, df.loc[counter_df][1], df.loc[counter_df][4], df.loc[counter_df][5],
                 df.loc[counter_df][6], df.loc[counter_df][8], df.loc[counter_df][9]]
    # 1. Checking the list has the same part number
    if len(elements) == 0:
        # Adding into the list
        elements.append(temp_list)
    # 2. Sperate the part number with - add addinto list; Adding all the total into the list
    else:
        for SingleElement in elements:  # i is inside array
            if Full_string == SingleElement[0]:
                # Sum
                # <1!---------- Adding more if more column   ---------->
                SingleElement[2] = df.loc[counter_df][4] + SingleElement[2]
                SingleElement[3] = df.loc[counter_df][5] + SingleElement[3]
                SingleElement[4] = df.loc[counter_df][6] + SingleElement[4]
                SingleElement[5] = df.loc[counter_df][8] + SingleElement[5]
                # <!------------Adding at here------------------------->
                SingleElement[6] = df.loc[counter_df][9] + SingleElement[6]
                # <!--------------------------------------------------->
                found = True
                break
            else:
                found = False

        if found == False:
            # Adding into 2D list
            elements.append(temp_list)
    return elements


def save_dictonary(Dict):
    #    save_file_name ="C:\\Users\\willlee\\Desktop\\PLANNER_INPUT_FORECAST.txt"
    f = open(save_file_name, "w")
    # Create First Row <!------------- Add more and add last column before "\n" if needed ----------->
    f.write(
        "Part Number" + "\t" + "Description" + "\t" + "Q2-FCST-Deduct" + "\t" + "SO total allocated" + "\t" + "SS1 total allocated" + "\t" + "BTG Q2" + "\t" + "SO + FC" + "\n")
    for key, value in Dict.items():
        Title = 'Planner_Code:{0};CCA/Module:{1}'.format(key[0], key[1])
        f.write(Title + "\n")
        f.write("-------------------------------------------\n")
        for OneDArray in value:
            for i in OneDArray:
                f.write(str(i) + "\t")
            f.write("\n")
        f.write("-------------------------------------------\n")
    f.close()


def write_into_template(moduleSelect, column_to_select):
    #    error_file = "C:\\Users\\willlee\\Desktop\\error_miss.txt"
    #    template_file_io_to_write ="C:\\Users\\willlee\\Desktop\\Cap_Ana_Controller_Q1_2020_SO+FC+25SS - Copy.xlsx"
    #    file_io_from_write = "C:\\Users\\willlee\\Desktop\\PLANNER_INPUT_FORECAST.txt"
    workbook = load_workbook(filename=template_file_io_to_write)
    worksheet = workbook["Data set"]
    if moduleSelect:
        # 1. Open file_io_from_write with column to select (part number and volume)
        with open(save_file_name, 'r') as fp:
            while True:
                chunk = fp.readline()
                split_build_descriptor = chunk.split(";")
                if spliiter_successfully_checked(split_build_descriptor):
                    # < !-------------Can change accordingly-------------------------------->
                    if split_build_descriptor[0].split(":")[1] == "CONT" and split_build_descriptor[1].split(":")[
                        1] == "Module\n":
                        # < !------------------------------------------------------------------->
                        # skipping next ---- line
                        chunk = fp.readline()
                        while True:
                            # <!--- From here it should be value ----->
                            WritingSuccessful = False  # <!-- Flag that using write into worksheet--->
                            chunk = fp.readline()
                            # Start splitting the \t
                            element = chunk.split("\t")
                            print(element)
                            # <! -------------- i and min_row should be same value and max_row should be depend on the value ----------->
                            i = 5
                            for value in worksheet.iter_rows(min_row=5, max_row=50, min_col=1, max_col=1,
                                                             values_only=True):
                                # Access the cell
                                print(value)
                                if value[0] != None:
                                    if value[0] == element[0]:
                                        #  <!------------------------------Can Change Accordingly -------------------------------------------->
                                        #  Write the value into the different Column and specifiy which column and row is writing  for both worksheet and array
                                        column_name = "C" + str(i)
                                        worksheet[column_name] = int(element[6].split('.')[0])
                                        # how about if worksheet no exsits, what should do?
                                        WritingSuccessful = True
                                i = i + 1
                            if not WritingSuccessful:
                                # <!----miss to write data into template, cannot find in template write into another file--->
                                build_String = ""
                                f = open(error_file, "a+")
                                for k in element:
                                    build_String = build_String + "\t" + k
                                f.write(build_String)
                                f.write("\n")
                                f.close()

                            if "--------" in chunk:
                                break
                else:
                    # No process if unsuccessful
                    pass
                if not chunk:
                    break
    else:
        print("No Module Defined!")
    # 2. Open template file and get the column  and write into and close
    workbook.save(filename=template_file_io_to_write)
    workbook.close()


def main():
    # 1. Get Every Sheet Name
    xl = pd.ExcelFile(file_name)
    print(xl.sheet_names)
    xl.close()
    # 2. Open Excel File Read with the Correct Sheet
    df = pd.read_excel(io=file_name, sheet_name=xl.sheet_names[3])
    print(df.head(5))
    # 3. Need to read first row for every column to use in group by
    # df.columns = df.columns.str.strip()
    # print (df.columns)
    # . Get the planner Code and Module level]
    df_grouping = df.groupby(['Planner Code', 'CCA/Module'])
    print(df_grouping)
    Dict = {}
    elements = []
    # elements.append([])
    for key, value in df_grouping.groups.items():
        print(key, value)
        temp_list = [value]
        for i in value:
            print(df.loc[i])
            print(type(df.loc[i]))
            # df.loc[i][0]
            elements = part_number_process(elements, df, i)

        Dict[key] = elements
        del elements
        elements = []

    # Iterate over key/value for debug purposes # Comment out if need to run faster
    for key, value in Dict.items():
        print(key, ':', value)
    save_dictonary(Dict)


if __name__ == "__main__":
    start_time = time.time()
    main()
    write_into_template(moduleSelect="CONT", column_to_select="A")
    print("%s seconds", time.time() - start_time)
